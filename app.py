
from flask import Flask, render_template, request, redirect, session, flash, send_file, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import sqlite3
import csv
import io
from openpyxl import load_workbook, Workbook
import os
import datetime
import zipfile
import smtplib
from email.message import EmailMessage

app = Flask(__name__)
app.secret_key = "CHANGE_THIS_SECRET_BEFORE_PRODUCTION"
DB = "atharvkart_v2.db"
UPLOAD = "static/uploads"

def now(): return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
def db():
    # SQLite connection configured for multi-request Flask usage.
    # WAL mode and busy_timeout prevent most "database is locked" errors.
    conn = sqlite3.connect(DB, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def safe_commit(conn, retries=5):
    import time
    for attempt in range(retries):
        try:
            conn.commit()
            return
        except sqlite3.OperationalError as e:
            if "locked" not in str(e).lower() or attempt == retries - 1:
                raise
            time.sleep(0.25 * (attempt + 1))


def init():
    c=db()
    c.executescript("""
    CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT,email TEXT UNIQUE,password TEXT,role TEXT DEFAULT 'customer',active INTEGER DEFAULT 1,phone TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS categories(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT UNIQUE,icon TEXT);
    CREATE TABLE IF NOT EXISTS products(id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT,description TEXT,price REAL,discount REAL DEFAULT 0,stock INTEGER DEFAULT 0,category_id INTEGER,image TEXT,seller_id INTEGER,active INTEGER DEFAULT 1,created_at TEXT);
    CREATE TABLE IF NOT EXISTS product_images(id INTEGER PRIMARY KEY AUTOINCREMENT,product_id INTEGER,image TEXT);
    CREATE TABLE IF NOT EXISTS cart(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,product_id INTEGER,qty INTEGER DEFAULT 1,UNIQUE(user_id,product_id));
    CREATE TABLE IF NOT EXISTS wishlist(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,product_id INTEGER,UNIQUE(user_id,product_id));
    CREATE TABLE IF NOT EXISTS addresses(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,full_name TEXT,phone TEXT,address TEXT,city TEXT,pincode TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS orders(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,address TEXT,total REAL,status TEXT DEFAULT 'Ordered',coupon_code TEXT,payment_mode TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS order_items(id INTEGER PRIMARY KEY AUTOINCREMENT,order_id INTEGER,product_id INTEGER,product_name TEXT,price REAL,qty INTEGER);
    CREATE TABLE IF NOT EXISTS coupons(id INTEGER PRIMARY KEY AUTOINCREMENT,code TEXT UNIQUE,discount INTEGER,active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS reviews(id INTEGER PRIMARY KEY AUTOINCREMENT,product_id INTEGER,user_id INTEGER,rating INTEGER,review TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS banners(id INTEGER PRIMARY KEY AUTOINCREMENT,title TEXT,subtitle TEXT,image TEXT,active INTEGER DEFAULT 1);
    CREATE TABLE IF NOT EXISTS settings(id INTEGER PRIMARY KEY AUTOINCREMENT,key TEXT UNIQUE,value TEXT);
    CREATE TABLE IF NOT EXISTS seller_requests(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,business_name TEXT,gst_no TEXT,status TEXT DEFAULT 'Pending',created_at TEXT);
    CREATE TABLE IF NOT EXISTS notifications(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,title TEXT,message TEXT,created_at TEXT,read_status INTEGER DEFAULT 0);
    CREATE TABLE IF NOT EXISTS sms_logs(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,order_id INTEGER,mobile TEXT,message TEXT,status TEXT,response TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS payments(id INTEGER PRIMARY KEY AUTOINCREMENT,order_id INTEGER,method TEXT,status TEXT,transaction_id TEXT,created_at TEXT);
    CREATE TABLE IF NOT EXISTS delivery_settings(id INTEGER PRIMARY KEY AUTOINCREMENT,min_order REAL DEFAULT 0,delivery_charge REAL DEFAULT 0,gst_percent REAL DEFAULT 18);
    CREATE TABLE IF NOT EXISTS order_finance(id INTEGER PRIMARY KEY AUTOINCREMENT,order_id INTEGER,seller_id INTEGER,product_value REAL,admin_margin REAL,seller_payable REAL,delivery_charge REAL,created_at TEXT);
    CREATE TABLE IF NOT EXISTS seller_wallet(id INTEGER PRIMARY KEY AUTOINCREMENT,seller_id INTEGER UNIQUE,balance REAL DEFAULT 0,total_sales REAL DEFAULT 0,total_margin_paid REAL DEFAULT 0,updated_at TEXT);
    CREATE TABLE IF NOT EXISTS audit_logs(id INTEGER PRIMARY KEY AUTOINCREMENT,user_id INTEGER,action TEXT,created_at TEXT);
    """)
    # Backward-compatible migration for existing V3.7/V3.8 databases.
    cols = [r[1] for r in c.execute("PRAGMA table_info(users)").fetchall()]
    if "phone" not in cols:
        c.execute("ALTER TABLE users ADD COLUMN phone TEXT")

    if not c.execute("SELECT 1 FROM users WHERE email='admin@atharvkart.com'").fetchone():
        for n,e,p,r in [
            ("Administrator","admin@atharvkart.com","admin123","admin"),
            ("Demo Seller","seller@atharvkart.com","seller123","seller")]:
            c.execute("INSERT INTO users(name,email,password,role,created_at) VALUES(?,?,?,?,?)",
                      (n,e,generate_password_hash(p),r,now()))
    for x in [("Mobiles","📱"),("Electronics","💻"),("Fashion","👕"),("Home & Kitchen","🏠"),("Books","📚"),("Beauty","💄")]:
        c.execute("INSERT OR IGNORE INTO categories(name,icon) VALUES(?,?)",x)
    for x in [("WELCOME10",10),("FESTIVE20",20)]:
        c.execute("INSERT OR IGNORE INTO coupons(code,discount) VALUES(?,?)",x)
    c.execute("INSERT OR IGNORE INTO banners(id,title,subtitle,active) VALUES(1,'Big Shopping Festival','Best offers on Electronics, Fashion and More',1)")
    for k,v in [('site_name','AtharvKart'),('support_email',''),('smtp_host','smtp.gmail.com'),('smtp_port','587'),('smtp_user',''),('smtp_password',''),('sms_enabled','0'),('sms_provider','MSG91'),('sms_authkey',''),('sms_sender_id',''),('sms_flow_id',''),('sms_country_code','91'),('delivery_charge','0'),('free_delivery_above','999')]:
        c.execute("INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)",(k,v))
    c.execute("INSERT OR IGNORE INTO delivery_settings(id,min_order,delivery_charge,gst_percent) VALUES(1,999,0,18)")
    safe_commit(c); c.close()

def login_required(): return bool(session.get("uid"))
def seller_access(): return session.get("role") in ("seller","admin")
def admin_access(): return session.get("role")=="admin"


def setting(key, default=""):
    c=db(); r=c.execute("SELECT value FROM settings WHERE key=?",(key,)).fetchone(); c.close()
    return r["value"] if r else default

def set_setting(key,value):
    c=db(); c.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",(key,str(value))); safe_commit(c); c.close()

def log_action(action):
    if session.get("uid"):
        c=db(); c.execute("INSERT INTO audit_logs(user_id,action,created_at) VALUES(?,?,?)",(session["uid"],action,now()));safe_commit(c);c.close()

def notify(user_id,title,message):
    c=db();c.execute("INSERT INTO notifications(user_id,title,message,created_at) VALUES(?,?,?,?)",(user_id,title,message,now()));safe_commit(c);c.close()

@app.context_processor
def notification_context():
    if not session.get("uid"):
        return {"unread_notifications": 0}
    try:
        c = db()
        count = c.execute("SELECT COUNT(*) FROM notifications WHERE user_id=? AND read_status=0", (session["uid"],)).fetchone()[0]
        c.close()
        return {"unread_notifications": count}
    except Exception:
        return {"unread_notifications": 0}

def send_email(to,subject,body):
    user=setting("smtp_user"); password=setting("smtp_password")
    if not user or not password or not to: return False
    try:
        msg=EmailMessage();msg["Subject"]=subject;msg["From"]=user;msg["To"]=to;msg.set_content(body)
        with smtplib.SMTP(setting("smtp_host","smtp.gmail.com"),int(setting("smtp_port","587"))) as s:
            s.starttls();s.login(user,password);s.send_message(msg)
        return True
    except Exception:
        return False

@app.before_request
def setup():
    os.makedirs(UPLOAD, exist_ok=True)
    # Database is initialized only once; do not perform schema writes on every request.
    if not getattr(app, "_db_ready", False):
        try:
            init()
            app._db_ready = True
        except sqlite3.OperationalError:
            # Another request may be completing initialization; retry on next request.
            pass

@app.route("/")
def home():
    c=db(); q=request.args.get("q","").strip(); cat=request.args.get("cat","")
    sql="""SELECT p.*,categories.name category,
        COALESCE((SELECT ROUND(AVG(rating),1) FROM reviews WHERE product_id=p.id),0) rating
        FROM products p LEFT JOIN categories ON categories.id=p.category_id WHERE p.active=1"""
    args=[]
    if q: sql+=" AND (p.name LIKE ? OR p.description LIKE ?)"; args += ["%"+q+"%","%"+q+"%"]
    if cat: sql+=" AND p.category_id=?"; args += [cat]
    sql+=" ORDER BY p.id DESC"
    products=c.execute(sql,args).fetchall()
    cats=c.execute("SELECT * FROM categories").fetchall()
    banners=c.execute("SELECT * FROM banners WHERE active=1 ORDER BY id DESC").fetchall()
    c.close()
    return render_template("home.html",products=products,cats=cats,banners=banners)

@app.route("/register",methods=["GET","POST"])
def register():
    if request.method=="POST":
        try:
            c=db()
            role=request.form.get("role","customer")
            if role=="seller": role="customer"
            uid=c.execute("INSERT INTO users(name,email,password,role,created_at) VALUES(?,?,?,?,?)",
                      (request.form["name"],request.form["email"],generate_password_hash(request.form["password"]),role,now())).lastrowid
            if request.form.get("seller_request")=="yes":
                c.execute("INSERT INTO seller_requests(user_id,business_name,gst_no,status,created_at) VALUES(?,?,?,?,?)",
                          (uid,request.form.get("business_name",""),request.form.get("gst_no",""),"Pending",now()))
            safe_commit(c);c.close();flash("Registration successful. Please login.");return redirect("/login")
        except Exception: flash("Email already registered")
    return render_template("register.html")

@app.route("/login",methods=["GET","POST"])
def login():
    if request.method=="POST":
        c=db();u=c.execute("SELECT * FROM users WHERE email=? AND active=1",(request.form["email"],)).fetchone();c.close()
        if u and check_password_hash(u["password"],request.form["password"]):
            session.update(uid=u["id"],name=u["name"],role=u["role"]);return redirect("/")
        flash("Invalid email or password")
    return render_template("login.html")

@app.route("/logout")
def logout(): session.clear();return redirect("/")

@app.route("/product/<int:id>",methods=["GET","POST"])
def product(id):
    c=db()
    if request.method=="POST":
        if not login_required(): return redirect("/login")
        c.execute("INSERT INTO reviews(product_id,user_id,rating,review,created_at) VALUES(?,?,?,?,?)",
                  (id,session["uid"],int(request.form["rating"]),request.form.get("review",""),now()))
        safe_commit(c);flash("Review added")
    p=c.execute("""SELECT p.*,categories.name category,COALESCE((SELECT ROUND(AVG(rating),1) FROM reviews WHERE product_id=p.id),0) rating
                   FROM products p LEFT JOIN categories ON categories.id=p.category_id WHERE p.id=?""",(id,)).fetchone()
    if p is None:
        c.close()
        flash("Product not found.")
        return redirect("/")
    imgs=c.execute("SELECT * FROM product_images WHERE product_id=?",(id,)).fetchall()
    reviews=c.execute("""SELECT reviews.*,users.name FROM reviews JOIN users ON users.id=reviews.user_id
                         WHERE product_id=? ORDER BY reviews.id DESC""",(id,)).fetchall()
    c.close()
    return render_template("product.html",p=p,imgs=imgs,reviews=reviews)

@app.route("/cart/add/<int:id>")
def cart_add(id):
    if not login_required(): return redirect("/login")
    c=db();x=c.execute("SELECT * FROM cart WHERE user_id=? AND product_id=?",(session["uid"],id)).fetchone()
    if x:c.execute("UPDATE cart SET qty=qty+1 WHERE id=?",(x["id"],))
    else:c.execute("INSERT INTO cart(user_id,product_id,qty) VALUES(?,?,1)",(session["uid"],id))
    safe_commit(c);c.close();flash("Product added to cart");return redirect(request.referrer or "/")

@app.route("/cart")
def cart():
    if not login_required(): return redirect("/login")
    c=db();rows=c.execute("""SELECT cart.*,products.name,products.price,products.discount,products.image
                             FROM cart JOIN products ON products.id=cart.product_id WHERE cart.user_id=?""",(session["uid"],)).fetchall()
    total=sum(r["price"]*(1-r["discount"]/100)*r["qty"] for r in rows);c.close()
    return render_template("cart.html",rows=rows,total=total)

@app.route("/cart/update/<int:id>/<action>")
def cart_update(id,action):
    c=db()
    if action=="plus":c.execute("UPDATE cart SET qty=qty+1 WHERE id=? AND user_id=?",(id,session["uid"]))
    if action=="minus":c.execute("UPDATE cart SET qty=MAX(1,qty-1) WHERE id=? AND user_id=?",(id,session["uid"]))
    if action=="remove":c.execute("DELETE FROM cart WHERE id=? AND user_id=?",(id,session["uid"]))
    safe_commit(c);c.close();return redirect("/cart")

@app.route("/wishlist/<int:id>")
def wishlist(id):
    if not login_required():return redirect("/login")
    c=db();c.execute("INSERT OR IGNORE INTO wishlist(user_id,product_id) VALUES(?,?)",(session["uid"],id));safe_commit(c);c.close();flash("Added to wishlist");return redirect(request.referrer or "/")

def send_sms(user_id, order_id, mobile, customer_name, total):
    """Send order-confirmation SMS through MSG91 Flow API when configured.
    Indian DLT template approval and sender/flow configuration are required.
    """
    mobile = (mobile or "").strip().replace("+", "")
    if mobile.startswith("0") and len(mobile) == 10:
        mobile = setting("sms_country_code", "91") + mobile
    elif len(mobile) == 10 and not mobile.startswith("91"):
        mobile = setting("sms_country_code", "91") + mobile
    enabled = setting("sms_enabled", "0") == "1"
    authkey = setting("sms_authkey", "")
    flow_id = setting("sms_flow_id", "")
    sender = setting("sms_sender_id", "")
    if not enabled or not authkey or not flow_id or not sender or len(mobile) < 10:
        return False
    msg = f"Dear {customer_name}, your AtharvKart order #{order_id} is confirmed. Total amount Rs.{total}. Thank you for shopping with us."
    payload = {
        "flow_id": flow_id,
        "sender": sender,
        "recipients": [{
            "mobiles": mobile,
            "VAR1": customer_name,
            "VAR2": str(order_id),
            "VAR3": f"{total:.2f}"
        }]
    }
    response_text = ""
    status = "Failed"
    try:
        req = urllib.request.Request(
            "https://control.msg91.com/api/v5/flow",
            data=json.dumps(payload).encode("utf-8"),
            headers={"accept":"application/json","authkey":authkey,"content-type":"application/json"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            response_text = resp.read().decode("utf-8", errors="replace")
            status = "Sent" if 200 <= resp.status < 300 else "Failed"
    except Exception as e:
        response_text = str(e)[:1000]
    try:
        c=db(); c.execute("INSERT INTO sms_logs(user_id,order_id,mobile,message,status,response,created_at) VALUES(?,?,?,?,?,?,?)",(user_id,order_id,mobile,msg,status,response_text,now())); safe_commit(c); c.close()
    except Exception:
        pass
    return status == "Sent"

@app.route("/wishlist")
def wishlist_page():
    if not login_required():return redirect("/login")
    c=db();rows=c.execute("""SELECT products.* FROM wishlist JOIN products ON products.id=wishlist.product_id WHERE wishlist.user_id=?""",(session["uid"],)).fetchall();c.close()
    return render_template("wishlist.html",rows=rows)

@app.route("/checkout",methods=["GET","POST"])
def checkout():
    if not login_required(): return redirect("/login")
    c=db(); rows=c.execute("""SELECT cart.*,products.name,products.price,products.discount,products.stock,products.seller_id
                             FROM cart JOIN products ON products.id=cart.product_id WHERE cart.user_id=?""",(session["uid"],)).fetchall()
    subtotal=round(sum(r["price"]*(1-r["discount"]/100)*r["qty"] for r in rows),2)
    discount=0; coupon=request.args.get("coupon","")
    if coupon:
        cp=c.execute("SELECT * FROM coupons WHERE code=? AND active=1",(coupon.upper(),)).fetchone()
        if cp: discount=round(subtotal*cp["discount"]/100,2)
    free_above=float(setting("free_delivery_above","999") or 999)
    delivery_charge=0 if subtotal>=free_above else float(setting("delivery_charge","0") or 0)
    # Customer pays only product net amount + delivery. Admin commission is never shown to customer.
    total=round(max(0,subtotal-discount)+delivery_charge,2)
    if request.method=="POST":
        coupon=request.form.get("coupon","").upper(); discount=0
        if coupon:
            cp=c.execute("SELECT * FROM coupons WHERE code=? AND active=1",(coupon,)).fetchone()
            if cp: discount=round(subtotal*cp["discount"]/100,2)
        free_above=float(setting("free_delivery_above","999") or 999)
        delivery_charge=0 if subtotal>=free_above else float(setting("delivery_charge","0") or 0)
        product_net=round(max(0,subtotal-discount),2)
        total=round(product_net+delivery_charge,2)
        customer_phone=request.form.get("phone","").strip()
        try: c.execute("UPDATE users SET phone=? WHERE id=?",(customer_phone,session["uid"]))
        except sqlite3.OperationalError: pass
        address=f"{request.form['full_name']} | {customer_phone} | {request.form['address']} | {request.form['city']} - {request.form['pincode']}"
        oid=c.execute("INSERT INTO orders(user_id,address,total,status,coupon_code,payment_mode,created_at) VALUES(?,?,?,?,?,?,?)",(session["uid"],address,total,"Ordered",coupon,request.form["payment_mode"],now())).lastrowid
        for r in rows:
            c.execute("INSERT INTO order_items(order_id,product_id,product_name,price,qty) VALUES(?,?,?,?,?)",(oid,r["product_id"],r["name"],r["price"],r["qty"]))
            c.execute("UPDATE products SET stock=MAX(0,stock-?) WHERE id=?",(r["qty"],r["product_id"]))
        # Internal seller settlement: 2% admin margin is deducted from seller product value only.
        seller_rows=c.execute("""SELECT p.seller_id,u.name seller_name,u.email seller_email,
                       SUM(oi.qty) item_qty,SUM((oi.price*(1-COALESCE(p.discount,0)/100))*oi.qty) item_amount,
                       GROUP_CONCAT(oi.product_name || ' x' || oi.qty, ', ') item_summary
                       FROM order_items oi JOIN products p ON p.id=oi.product_id LEFT JOIN users u ON u.id=p.seller_id
                       WHERE oi.order_id=? AND p.seller_id IS NOT NULL GROUP BY p.seller_id""",(oid,)).fetchall()
        for sr in seller_rows:
            seller_id=sr["seller_id"]; amount=round(float(sr["item_amount"] or 0),2); margin=round(amount*0.02,2); payable=round(amount-margin,2)
            c.execute("INSERT INTO order_finance(order_id,seller_id,product_value,admin_margin,seller_payable,delivery_charge,created_at) VALUES(?,?,?,?,?,?,?)",(oid,seller_id,amount,margin,payable,0,now()))
            c.execute("INSERT INTO seller_wallet(seller_id,balance,total_sales,total_margin_paid,updated_at) VALUES(?,?,?,?,?) ON CONFLICT(seller_id) DO UPDATE SET balance=balance+excluded.balance,total_sales=total_sales+excluded.total_sales,total_margin_paid=total_margin_paid+excluded.total_margin_paid,updated_at=excluded.updated_at",(seller_id,payable,amount,margin,now()))
        c.execute("INSERT INTO payments(order_id,method,status,transaction_id,created_at) VALUES(?,?,?,?,?)",(oid,request.form["payment_mode"],"Pending" if request.form["payment_mode"]!="Cash on Delivery" else "COD",request.form.get("transaction_id","DEMO-"+str(oid)),now()))
        email_row=c.execute("SELECT email FROM users WHERE id=?",(session["uid"],)).fetchone(); c.execute("DELETE FROM cart WHERE user_id=?",(session["uid"],)); safe_commit(c); c.close()
        notify(session["uid"],"Order Placed",f"Your order #{oid} has been placed successfully. Total ₹{total}")
        send_email(email_row["email"] if email_row else "",f"AtharvKart Order #{oid}",f"Your order has been placed successfully. Total amount: ₹{total}")
        send_sms(session["uid"],oid,customer_phone,request.form.get("full_name",session.get("name","Customer")),total)
        for sr in seller_rows:
            msg=f"New order #{oid} received from {session.get('name','Customer')}. Items: {sr['item_summary'] or 'New product order'}. Please process the order."
            notify(sr["seller_id"],"🛒 New Order Received",msg); send_email(sr["seller_email"] or "",f"New AtharvKart Order #{oid}",msg)
        log_action(f"Placed order #{oid}"); flash(f"Order #{oid} placed successfully"); return redirect("/orders")
    c.close(); return render_template("checkout.html",rows=rows,subtotal=subtotal,discount=discount,delivery_charge=delivery_charge,gst=0,total=total)

@app.route("/orders")
def orders():
    if not login_required():return redirect("/login")
    c=db();rows=c.execute("SELECT * FROM orders WHERE user_id=? ORDER BY id DESC",(session["uid"],)).fetchall();c.close()
    return render_template("orders.html",rows=rows)

@app.route("/invoice/<int:id>")
def invoice(id):
    if not login_required():return redirect("/login")
    c=db();o=c.execute("SELECT * FROM orders WHERE id=? AND user_id=?",(id,session["uid"])).fetchone()
    items=c.execute("SELECT * FROM order_items WHERE order_id=?",(id,)).fetchall();c.close()
    return render_template("invoice.html",o=o,items=items)


@app.route("/product/edit/<int:id>",methods=["GET","POST"])
def product_edit(id):
    if not login_required() or not seller_access(): return redirect("/")
    c=db(); p=c.execute("SELECT * FROM products WHERE id=?",(id,)).fetchone()
    if not p: c.close(); return redirect("/seller")
    if request.method=="POST":
        f=request.form
        c.execute("UPDATE products SET name=?,description=?,price=?,discount=?,stock=?,category_id=?,active=? WHERE id=?",
                  (f["name"],f["description"],float(f["price"]),float(f.get("discount",0)),int(f["stock"]),int(f["category_id"]),int(f.get("active",1)),id))
        safe_commit(c);c.close();flash("Product updated");log_action(f"Updated product #{id}");return redirect("/seller")
    cats=c.execute("SELECT * FROM categories").fetchall();c.close()
    return render_template("product_edit.html",p=p,cats=cats)

@app.route("/product/delete/<int:id>")
def product_delete(id):
    if not login_required() or not admin_access(): return redirect("/")
    c=db();c.execute("DELETE FROM products WHERE id=?",(id,));safe_commit(c);c.close();flash("Product deleted");log_action(f"Deleted product #{id}");return redirect("/admin")

@app.route("/seller/apply",methods=["GET","POST"])
def seller_apply():
    if not login_required(): return redirect("/login")
    if request.method=="POST":
        c=db();c.execute("INSERT INTO seller_requests(user_id,business_name,gst_no,status,created_at) VALUES(?,?,?,?,?)",
                         (session["uid"],request.form["business_name"],request.form.get("gst_no",""),"Pending",now()));safe_commit(c);c.close()
        flash("Seller approval request submitted");return redirect("/")
    return render_template("seller_apply.html")

@app.route("/notifications")
def notifications():
    if not login_required(): return redirect("/login")
    c=db(); rows=c.execute("SELECT * FROM notifications WHERE user_id=? ORDER BY id DESC",(session["uid"],)).fetchall()
    c.execute("UPDATE notifications SET read_status=1 WHERE user_id=?",(session["uid"],));safe_commit(c);c.close()
    return render_template("notifications.html",rows=rows)

@app.route("/settings",methods=["GET","POST"])
def settings():
    if not login_required() or not admin_access(): return redirect("/")
    keys=["site_name","support_email","smtp_host","smtp_port","smtp_user","smtp_password","sms_enabled","sms_provider","sms_authkey","sms_sender_id","sms_flow_id","sms_country_code","delivery_charge","free_delivery_above"]
    if request.method=="POST":
        for k in keys:
            if k in request.form: set_setting(k,request.form[k])
        flash("Settings saved successfully");log_action("Updated system settings")
    values={k:setting(k) for k in keys}
    return render_template("settings.html",values=values)

@app.route("/admin/sms-logs")
def admin_sms_logs():
    if not login_required() or not admin_access(): return redirect("/")
    c=db(); rows=c.execute("SELECT s.*,u.name AS customer_name FROM sms_logs s LEFT JOIN users u ON u.id=s.user_id ORDER BY s.id DESC LIMIT 200").fetchall(); c.close()
    return render_template("sms_logs.html", rows=rows)

@app.route("/admin/sellers")
def admin_sellers():
    if not login_required() or not admin_access(): return redirect("/")
    c=db();rows=c.execute("""SELECT seller_requests.*,users.name,users.email FROM seller_requests
                             JOIN users ON users.id=seller_requests.user_id ORDER BY seller_requests.id DESC""").fetchall();c.close()
    return render_template("seller_requests.html",rows=rows)

@app.route("/admin/seller/<int:id>/<action>")
def admin_seller_action(id,action):
    if not login_required() or not admin_access(): return redirect("/")
    c=db();r=c.execute("SELECT * FROM seller_requests WHERE id=?",(id,)).fetchone()
    if r:
        status="Approved" if action=="approve" else "Rejected"
        c.execute("UPDATE seller_requests SET status=? WHERE id=?",(status,id))
        if action=="approve": c.execute("UPDATE users SET role='seller' WHERE id=?",(r["user_id"],))
        notify(r["user_id"],"Seller Request "+status,f"Your seller account request has been {status}.")
        safe_commit(c)
    c.close();return redirect("/admin/sellers")

@app.route("/admin/finance")
def admin_finance():
    if not login_required() or not admin_access(): return redirect("/")
    c=db(); rows=c.execute("""SELECT f.*,u.name seller_name,o.user_id,o.status,o.created_at order_date
        FROM order_finance f JOIN users u ON u.id=f.seller_id JOIN orders o ON o.id=f.order_id
        ORDER BY f.id DESC LIMIT 500""").fetchall()
    totals=c.execute("SELECT COALESCE(SUM(product_value),0) product_value,COALESCE(SUM(admin_margin),0) admin_margin,COALESCE(SUM(seller_payable),0) seller_payable FROM order_finance").fetchone(); c.close()
    return render_template("admin_finance.html",rows=rows,totals=totals)

@app.route("/reports")
def reports():
    if not login_required() or not admin_access(): return redirect("/")
    c=db()
    low=c.execute("SELECT * FROM products WHERE stock<=5 ORDER BY stock ASC").fetchall()
    sales=c.execute("""SELECT DATE(created_at) day,ROUND(SUM(total),2) total,COUNT(*) orders
                       FROM orders WHERE status!='Cancelled' GROUP BY DATE(created_at) ORDER BY day DESC LIMIT 30""").fetchall()
    top=c.execute("""SELECT product_name,SUM(qty) qty,SUM(price*qty) amount FROM order_items
                     GROUP BY product_name ORDER BY qty DESC LIMIT 10""").fetchall()
    c.close();return render_template("reports.html",low=low,sales=sales,top=top)

@app.route("/seller",methods=["GET","POST"])
def seller():
    if not login_required() or not seller_access(): return redirect("/")
    c=db()
    if request.method=="POST":
        names=request.form.getlist("name"); descs=request.form.getlist("description"); prices=request.form.getlist("price"); discounts=request.form.getlist("discount"); stocks=request.form.getlist("stock"); cats=request.form.getlist("category_id"); imgs=request.files.getlist("image")
        count=0
        for idx,name in enumerate(names):
            if not name.strip(): continue
            img=imgs[idx] if idx < len(imgs) else None; filename=""
            if img and img.filename:
                filename=secure_filename(datetime.datetime.now().strftime("%Y%m%d%H%M%S_")+img.filename); img.save(os.path.join(UPLOAD,filename))
            c.execute("INSERT INTO products(name,description,price,discount,stock,category_id,image,seller_id,created_at) VALUES(?,?,?,?,?,?,?,?,?)",(name.strip(),descs[idx] if idx<len(descs) else "",float(prices[idx]),float(discounts[idx] or 0),int(stocks[idx]),int(cats[idx]),filename,session["uid"],now())); count+=1
        safe_commit(c); flash(f"{count} product(s) added successfully")
    cats=c.execute("SELECT * FROM categories").fetchall(); products=c.execute("SELECT p.*,categories.name category FROM products p LEFT JOIN categories ON categories.id=p.category_id WHERE p.seller_id=? ORDER BY p.id DESC",(session["uid"],)).fetchall(); sales=c.execute("SELECT COALESCE(SUM(seller_payable),0) FROM order_finance WHERE seller_id=?",(session["uid"],)).fetchone()[0]; c.close()
    return render_template("seller.html",cats=cats,products=products,sales=sales)

@app.route("/admin",methods=["GET","POST"])
def admin():
    if not login_required() or not admin_access():return redirect("/")
    c=db()
    if request.method=="POST":
        mode=request.form.get("mode")
        if mode=="category":c.execute("INSERT OR IGNORE INTO categories(name,icon) VALUES(?,?)",(request.form["name"],request.form.get("icon","🛍️")))
        elif mode=="coupon":c.execute("INSERT OR IGNORE INTO coupons(code,discount,active) VALUES(?,?,1)",(request.form["code"].upper(),int(request.form["discount"])))
        elif mode=="banner":c.execute("INSERT INTO banners(title,subtitle,active) VALUES(?,?,1)",(request.form["title"],request.form.get("subtitle","")))
        elif mode=="status":c.execute("UPDATE orders SET status=? WHERE id=?",(request.form["status"],request.form["order_id"]))
        safe_commit(c);flash("Updated successfully")
    stats={"Users":c.execute("SELECT COUNT(*) FROM users").fetchone()[0],"Products":c.execute("SELECT COUNT(*) FROM products").fetchone()[0],"Orders":c.execute("SELECT COUNT(*) FROM orders").fetchone()[0],"Sales":round(c.execute("SELECT COALESCE(SUM(total),0) FROM orders WHERE status!='Cancelled'").fetchone()[0],2)}
    orders=c.execute("SELECT orders.*,users.name FROM orders JOIN users ON users.id=orders.user_id ORDER BY orders.id DESC LIMIT 50").fetchall()
    coupons=c.execute("SELECT * FROM coupons ORDER BY id DESC").fetchall();c.close()
    return render_template("admin.html",stats=stats,orders=orders,coupons=coupons)

@app.route("/backup")
def backup():
    if not admin_access():return redirect("/")
    name="ATHARVKART_BACKUP_"+datetime.datetime.now().strftime("%Y%m%d_%H%M%S")+".zip"
    with zipfile.ZipFile(name,"w",zipfile.ZIP_DEFLATED) as z:
        if os.path.exists(DB):z.write(DB)
        for root,_,files in os.walk(UPLOAD):
            for f in files:z.write(os.path.join(root,f),arcname=os.path.join("uploads",f))
    return send_file(name,as_attachment=True)


# -------------------- EXTRA WORKING MENU ROUTES --------------------

@app.route("/dashboard")
def dashboard():
    if not login_required():
        return redirect("/login")
    c = db()
    if admin_access():
        stats = {
            "Users": c.execute("SELECT COUNT(*) FROM users").fetchone()[0],
            "Products": c.execute("SELECT COUNT(*) FROM products").fetchone()[0],
            "Orders": c.execute("SELECT COUNT(*) FROM orders").fetchone()[0],
            "Sales": round(c.execute("SELECT COALESCE(SUM(total),0) FROM orders WHERE status!='Cancelled'").fetchone()[0], 2)
        }
    elif session.get("role") == "seller":
        stats = {
            "My Products": c.execute("SELECT COUNT(*) FROM products WHERE seller_id=?", (session["uid"],)).fetchone()[0],
            "Low Stock": c.execute("SELECT COUNT(*) FROM products WHERE seller_id=? AND stock<=5", (session["uid"],)).fetchone()[0],
            "My Orders": c.execute("""SELECT COUNT(DISTINCT oi.order_id)
                                      FROM order_items oi
                                      JOIN products p ON p.id=oi.product_id
                                      WHERE p.seller_id=?""", (session["uid"],)).fetchone()[0]
        }
    else:
        stats = {
            "My Orders": c.execute("SELECT COUNT(*) FROM orders WHERE user_id=?", (session["uid"],)).fetchone()[0],
            "Cart Items": c.execute("SELECT COALESCE(SUM(qty),0) FROM cart WHERE user_id=?", (session["uid"],)).fetchone()[0],
            "Wishlist": c.execute("SELECT COUNT(*) FROM wishlist WHERE user_id=?", (session["uid"],)).fetchone()[0]
        }
    c.close()
    return render_template("dashboard.html", stats=stats)

@app.route("/seller/orders")
def seller_orders():
    if not login_required() or not seller_access():
        return redirect("/")
    c = db()
    rows = c.execute("""
        SELECT o.id AS order_id, o.created_at, o.status, o.payment_mode,
               u.name AS customer_name, u.email AS customer_email,
               oi.product_name, oi.price, oi.qty,
               ROUND(oi.price * oi.qty, 2) AS item_total
        FROM orders o
        JOIN users u ON u.id=o.user_id
        JOIN order_items oi ON oi.order_id=o.id
        JOIN products p ON p.id=oi.product_id
        WHERE p.seller_id=?
        ORDER BY o.id DESC, oi.id DESC
    """, (session["uid"],)).fetchall()
    c.close()
    return render_template("seller_orders.html", rows=rows)


@app.route("/seller/products")
def seller_products():
    if not login_required() or not seller_access():
        return redirect("/")
    return redirect("/seller")

@app.route("/admin/products")
def admin_products():
    if not login_required() or not admin_access():
        return redirect("/")
    c = db()
    rows = c.execute("""SELECT p.*, categories.name AS category, users.name AS seller_name
                        FROM products p
                        LEFT JOIN categories ON categories.id=p.category_id
                        LEFT JOIN users ON users.id=p.seller_id
                        ORDER BY p.id DESC""").fetchall()
    c.close()
    return render_template("admin_products.html", rows=rows)

@app.route("/admin/categories", methods=["GET","POST"])
def admin_categories():
    if not login_required() or not admin_access():
        return redirect("/")
    c = db()
    if request.method == "POST":
        action = request.form.get("action", "add")
        if action == "add":
            c.execute("INSERT OR IGNORE INTO categories(name,icon) VALUES(?,?)",
                      (request.form["name"], request.form.get("icon","🛍️")))
            flash("Category added successfully")
        elif action == "delete":
            c.execute("DELETE FROM categories WHERE id=?", (request.form["id"],))
            flash("Category deleted successfully")
        safe_commit(c)
    rows = c.execute("SELECT * FROM categories ORDER BY id DESC").fetchall()
    c.close()
    return render_template("admin_categories.html", rows=rows)

@app.route("/admin/coupons", methods=["GET","POST"])
def admin_coupons():
    if not login_required() or not admin_access():
        return redirect("/")
    c = db()
    if request.method == "POST":
        action = request.form.get("action", "add")
        if action == "add":
            c.execute("INSERT OR IGNORE INTO coupons(code,discount,active) VALUES(?,?,1)",
                      (request.form["code"].upper(), int(request.form["discount"])))
            flash("Coupon added successfully")
        elif action == "toggle":
            c.execute("UPDATE coupons SET active=CASE WHEN active=1 THEN 0 ELSE 1 END WHERE id=?",
                      (request.form["id"],))
            flash("Coupon status updated")
        elif action == "delete":
            c.execute("DELETE FROM coupons WHERE id=?", (request.form["id"],))
            flash("Coupon deleted")
        safe_commit(c)
    rows = c.execute("SELECT * FROM coupons ORDER BY id DESC").fetchall()
    c.close()
    return render_template("admin_coupons.html", rows=rows)

@app.route("/admin/banners", methods=["GET","POST"])
def admin_banners():
    if not login_required() or not admin_access():
        return redirect("/")
    c = db()
    if request.method == "POST":
        action = request.form.get("action", "add")
        if action == "add":
            image = ""
            img = request.files.get("image")
            if img and img.filename:
                image = secure_filename(datetime.datetime.now().strftime("%Y%m%d%H%M%S_") + img.filename)
                img.save(os.path.join(UPLOAD, image))
            c.execute("INSERT INTO banners(title,subtitle,image,active) VALUES(?,?,?,1)",
                      (request.form["title"], request.form.get("subtitle",""), image))
            flash("Banner added successfully")
        elif action == "toggle":
            c.execute("UPDATE banners SET active=CASE WHEN active=1 THEN 0 ELSE 1 END WHERE id=?",
                      (request.form["id"],))
            flash("Banner status updated")
        elif action == "delete":
            c.execute("DELETE FROM banners WHERE id=?", (request.form["id"],))
            flash("Banner deleted")
        safe_commit(c)
    rows = c.execute("SELECT * FROM banners ORDER BY id DESC").fetchall()
    c.close()
    return render_template("admin_banners.html", rows=rows)

@app.route("/admin/orders", methods=["GET","POST"])
def admin_orders():
    if not login_required() or not admin_access():
        return redirect("/")
    c = db()
    if request.method == "POST":
        c.execute("UPDATE orders SET status=? WHERE id=?",
                  (request.form["status"], request.form["order_id"]))
        safe_commit(c)
        flash("Order status updated successfully")
    rows = c.execute("""SELECT orders.*, users.name, users.email
                        FROM orders JOIN users ON users.id=orders.user_id
                        ORDER BY orders.id DESC""").fetchall()
    c.close()
    return render_template("admin_orders.html", rows=rows)



# -------------------- ADVANCED DIGITAL ANALYTICS --------------------
@app.route("/analytics")
def analytics():
    if not login_required():
        return redirect("/login")
    c = db()

    if admin_access():
        sales_rows = c.execute("""
            SELECT substr(created_at,1,10) d, ROUND(SUM(total),2) total
            FROM orders WHERE status!='Cancelled'
            GROUP BY substr(created_at,1,10)
            ORDER BY d DESC LIMIT 7
        """).fetchall()
        order_status = c.execute("""
            SELECT status, COUNT(*) total FROM orders GROUP BY status ORDER BY total DESC
        """).fetchall()
        low_stock = c.execute("""
            SELECT name, stock FROM products WHERE stock<=5 ORDER BY stock ASC LIMIT 10
        """).fetchall()
        top_products = c.execute("""
            SELECT p.name, SUM(oi.qty) qty
            FROM order_items oi JOIN products p ON p.id=oi.product_id
            GROUP BY p.id ORDER BY qty DESC LIMIT 10
        """).fetchall()
    elif session.get("role") == "seller":
        sales_rows = c.execute("""
            SELECT substr(o.created_at,1,10) d, ROUND(SUM(oi.qty*oi.price),2) total
            FROM order_items oi
            JOIN orders o ON o.id=oi.order_id
            JOIN products p ON p.id=oi.product_id
            WHERE p.seller_id=? AND o.status!='Cancelled'
            GROUP BY substr(o.created_at,1,10)
            ORDER BY d DESC LIMIT 7
        """,(session["uid"],)).fetchall()
        order_status = c.execute("""
            SELECT o.status, COUNT(DISTINCT o.id) total
            FROM orders o JOIN order_items oi ON oi.order_id=o.id
            JOIN products p ON p.id=oi.product_id
            WHERE p.seller_id=?
            GROUP BY o.status ORDER BY total DESC
        """,(session["uid"],)).fetchall()
        low_stock = c.execute("""
            SELECT name, stock FROM products WHERE seller_id=? AND stock<=5 ORDER BY stock ASC LIMIT 10
        """,(session["uid"],)).fetchall()
        top_products = c.execute("""
            SELECT p.name, SUM(oi.qty) qty
            FROM order_items oi JOIN products p ON p.id=oi.product_id
            WHERE p.seller_id=?
            GROUP BY p.id ORDER BY qty DESC LIMIT 10
        """,(session["uid"],)).fetchall()
    else:
        sales_rows = c.execute("""
            SELECT substr(created_at,1,10) d, ROUND(SUM(total),2) total
            FROM orders WHERE user_id=? GROUP BY substr(created_at,1,10)
            ORDER BY d DESC LIMIT 7
        """,(session["uid"],)).fetchall()
        order_status = c.execute("""
            SELECT status, COUNT(*) total FROM orders WHERE user_id=? GROUP BY status
        """,(session["uid"],)).fetchall()
        low_stock = []
        top_products = []

    c.close()
    return render_template(
        "analytics.html",
        sales_rows=list(reversed(sales_rows)),
        order_status=order_status,
        low_stock=low_stock,
        top_products=top_products
    )


# -------------------- PRODUCT BULK UPLOAD --------------------
@app.route("/admin/bulk-product-upload", methods=["GET", "POST"])
def bulk_product_upload():
    if not login_required() or not admin_access():
        return redirect("/")

    summary = None
    errors = []
    if request.method == "POST":
        file = request.files.get("file")
        update_existing = request.form.get("update_existing") == "1"

        if not file or not file.filename:
            flash("Please select an Excel or CSV file.")
            return redirect(request.url)

        filename = secure_filename(file.filename.lower())
        rows = []

        try:
            if filename.endswith(".csv"):
                raw = file.read().decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(raw))
                rows = list(reader)
            elif filename.endswith(".xlsx"):
                wb = load_workbook(file, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
                for values in ws.iter_rows(min_row=2, values_only=True):
                    row = {}
                    for i, value in enumerate(values):
                        if i < len(headers):
                            row[headers[i]] = value
                    if any(v not in (None, "") for v in values):
                        rows.append(row)
            else:
                flash("Only .xlsx and .csv files are supported.")
                return redirect(request.url)
        except Exception as e:
            flash(f"Unable to read file: {e}")
            return redirect(request.url)

        c = db()
        inserted = updated = skipped = 0

        def get_value(row, *names):
            for name in names:
                if name in row and row[name] not in (None, ""):
                    return row[name]
            # case-insensitive fallback
            low = {str(k).strip().lower(): v for k, v in row.items()}
            for name in names:
                v = low.get(name.lower())
                if v not in (None, ""):
                    return v
            return ""

        try:
            for idx, row in enumerate(rows, start=2):
                try:
                    name = str(get_value(row, "Product Name", "name", "product_name")).strip()
                    category_name = str(get_value(row, "Category", "category")).strip()
                    price = float(get_value(row, "Price", "price") or 0)
                    mrp = float(get_value(row, "MRP", "mrp") or price)
                    stock = int(float(get_value(row, "Stock", "stock") or 0))
                    description = str(get_value(row, "Description", "description") or "").strip()
                    image = str(get_value(row, "Image", "Image URL", "image", "image_url") or "").strip()

                    if not name:
                        raise ValueError("Product Name is required")
                    if price < 0 or mrp < 0 or stock < 0:
                        raise ValueError("Price, MRP and Stock cannot be negative")

                    category_id = None
                    if category_name:
                        cat = c.execute("SELECT id FROM categories WHERE LOWER(name)=LOWER(?)", (category_name,)).fetchone()
                        if not cat:
                            c.execute("INSERT INTO categories(name,icon) VALUES(?,?)", (category_name, "🛍️"))
                            safe_commit(c)
                            cat = c.execute("SELECT id FROM categories WHERE LOWER(name)=LOWER(?)", (category_name,)).fetchone()
                        category_id = cat[0]

                    existing = c.execute("SELECT id FROM products WHERE LOWER(name)=LOWER(?)", (name,)).fetchone()
                    if existing:
                        if update_existing:
                            c.execute("""UPDATE products
                                         SET category_id=?, price=?, mrp=?, stock=?, description=?, image=?
                                         WHERE id=?""",
                                      (category_id, price, mrp, stock, description, image, existing[0]))
                            updated += 1
                        else:
                            skipped += 1
                            errors.append(f"Row {idx}: Duplicate product skipped - {name}")
                    else:
                        # seller_id NULL makes the product admin-owned.
                        c.execute("""INSERT INTO products(name,category_id,price,mrp,stock,description,image,seller_id)
                                     VALUES(?,?,?,?,?,?,?,NULL)""",
                                  (name, category_id, price, mrp, stock, description, image))
                        inserted += 1
                except Exception as e:
                    skipped += 1
                    errors.append(f"Row {idx}: {e}")

            safe_commit(c)
            summary = {
                "total": len(rows),
                "inserted": inserted,
                "updated": updated,
                "skipped": skipped
            }
        except Exception as e:
            c.rollback()
            flash(f"Bulk upload failed: {e}")
        finally:
            c.close()

    return render_template("bulk_product_upload.html", summary=summary, errors=errors)


@app.route("/admin/bulk-product-template")
def bulk_product_template():
    if not login_required() or not admin_access():
        return redirect("/")

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    headers = ["Product Name", "Category", "Price", "MRP", "Stock", "Description", "Image URL"]
    ws.append(headers)
    ws.append(["Sample Mobile Cover", "Electronics", 299, 499, 50, "Premium quality mobile cover", ""])
    ws.append(["Sample Laptop Bag", "Accessories", 999, 1499, 20, "Water resistant laptop bag", ""])

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="ATHARVKART_Product_Bulk_Upload_Template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__=="__main__":
    os.makedirs(UPLOAD,exist_ok=True);init()
    # Safe database migration for older V3 installations
    try:
        _c=db()
        cols=[r[1] for r in _c.execute("PRAGMA table_info(banners)").fetchall()]
        if "image" not in cols:
            _c.execute("ALTER TABLE banners ADD COLUMN image TEXT DEFAULT ''")
            _safe_commit(c)
        _c.close()
    except Exception:
        pass
    app.run(host="127.0.0.1",port=5000,debug=False,threaded=True,use_reloader=False)
